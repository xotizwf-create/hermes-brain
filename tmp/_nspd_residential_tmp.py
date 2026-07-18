#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Armed: for each target parcel, find the NEAREST RESIDENTIAL object in each of the 8
compass sectors — considering BOTH residential land parcels (by ВРИ) AND residential
BUILDINGS (by назначение/purpose: жилой дом / многоквартирный). Picks the single nearest
per sector across both layers, plus the overall nearest. One .xlsx, a sheet per target.

Needs a RU residential IP (VPN OFF). Reuses tiled complete-search from
scripts/nspd_parcels_local.py (defeats the silent ~300 cap).
"""
import importlib.util, json, os, sys, time

HERE = os.path.dirname(os.path.abspath(__file__))
spec = importlib.util.spec_from_file_location(
    "npl", os.path.join(HERE, "scripts", "nspd_parcels_local.py"))
npl = importlib.util.module_from_spec(spec); spec.loader.exec_module(npl)

from shapely.ops import transform as shp_transform

RADIUS = 1200.0
DEADLINE_MIN = 30
SECTORS = npl.SECTORS

TARGETS = [
    ("18:26:041310:31", "Склад Ижевск (Нагорная 3)"),
    ("18:30:000148:8",  "Промплощадка 2 Сарапул (Красный проезд 1)"),
]
OUTXLSX = os.path.join(HERE, "Ближайшая_жилая_застройка_по_сторонам.xlsx")
STATUS = os.path.join(HERE, "_nspd_resi_status.json")

# land-parcel residential markers (ВРИ), with hard exclusions
RESI_VRI = ["жил", "ижс", "индивидуальн", "малоэтаж", "среднеэтаж", "многоэтаж",
            "многоквартир", "блокированн", "приусадебн", "личного подсобного", "лпх", "усадеб"]
EXCLUDE_VRI = ["гараж", "стоянк", "эксплуатационные служб", "рэу", "прэо", "нежил",
               "промышл", "склад", "производ", "коммунальн", "огородн", "садовод",
               "сельскохоз", "торгов", "офис", "общественн", "магазин"]


def log(*a):
    print(*a, file=sys.stderr, flush=True)


def land_vri(o):
    return (o.get("permitted_use_established_by_document") or o.get("land_record_category_type") or "")


def is_resi_land(o):
    v = land_vri(o).lower()
    if not v:
        return False
    if any(k in v for k in EXCLUDE_VRI):
        return False
    return any(k in v for k in RESI_VRI)


def is_resi_building(o):
    p = (o.get("purpose") or "").lower()
    if not p:
        # fall back to record type / name if purpose is blank
        p = ((o.get("build_record_type_value") or "") + " " + (o.get("building_name") or "")).lower()
    if not p:
        return False
    if "нежил" in p:
        return False
    return ("жил" in p) or ("многоквартир" in p)


def enum_layer(nspd, layer, buf_native, target):
    def cad_key(f):
        o = npl.opt(f)
        return o.get("cad_num") or o.get("cad_number") or getattr(f, "cadastral_number", None)
    return npl.collect_complete(nspd, layer, buf_native, cad_key)


def find_for_target(nspd, target):
    central = nspd.find_landplot(target)
    if central is None:
        raise RuntimeError(f"target {target} not found (RU IP active?)")
    cg = npl.geom_of(central)
    to_utm, from_utm, utm, src = npl.detect_to_utm(cg)
    cen_utm = shp_transform(to_utm, cg)
    cen_centroid = cen_utm.centroid
    buf_native = shp_transform(from_utm, cen_utm.buffer(RADIUS))

    candidates = []  # (kind_label, vri_text, feature)
    plots = enum_layer(nspd, npl.Layer36048Feature, buf_native, target)
    log(f"  {target}: {len(plots)} land parcels in buffer")
    for f in plots:
        o = npl.opt(f)
        if is_resi_land(o):
            candidates.append(("Земельный участок (жилой ВРИ)", land_vri(o), f))
    blds = enum_layer(nspd, npl.Layer36049Feature, buf_native, target)
    log(f"  {target}: {len(blds)} buildings in buffer")
    for f in blds:
        o = npl.opt(f)
        if is_resi_building(o):
            purpose = o.get("purpose") or o.get("build_record_type_value") or "жилое"
            candidates.append((f"Здание ({purpose})", purpose, f))
    log(f"  {target}: {len(candidates)} residential candidates (parcels+buildings)")

    best = {}      # sector -> row
    overall = None
    for kind, vri, f in candidates:
        o = npl.opt(f)
        cad = o.get("cad_num") or o.get("cad_number")
        if not cad or cad == target:
            continue
        g = npl.geom_of(f)
        if g is None:
            continue
        try:
            g_utm = shp_transform(to_utm, g)
        except Exception:
            continue
        dist = cen_utm.distance(g_utm)
        if dist > RADIUS:
            continue
        sec = npl.bearing_sector(cen_centroid, g_utm.centroid)
        area = (o.get("specified_area") or o.get("declared_area") or o.get("area")
                or o.get("build_record_area") or "")
        row = {"sector": sec, "kind": kind, "cad": cad,
               "addr": o.get("readable_address") or "", "vri": vri,
               "dist": round(dist, 1), "area": area}
        if sec not in best or dist < best[sec]["dist"]:
            best[sec] = row
        if overall is None or dist < overall["dist"]:
            overall = row
    return best, overall


def write_xlsx(per_target):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    hl = PatternFill("solid", fgColor="FFF2CC")
    for target, title, best, overall in per_target:
        ws = wb.create_sheet((title[:28] or target))
        ws.append(["Сторона света", "Тип объекта", "Кад. номер", "Адрес",
                   "ВРИ / Назначение", "Расстояние, м", "Площадь, кв. м"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for sec in SECTORS:
            r = best.get(sec)
            if r:
                ws.append([sec, r["kind"], r["cad"], r["addr"], r["vri"], r["dist"], r["area"]])
            else:
                ws.append([sec, "— нет жилья в радиусе 1200 м", "", "", "", "", ""])
        ws.append([])
        if overall:
            row = ws.max_row + 1
            ws.append(["★ САМОЕ БЛИЖАЙШЕЕ", overall["kind"], overall["cad"], overall["addr"],
                       overall["vri"], overall["dist"], overall["area"]])
            for c in ws[ws.max_row]:
                c.fill = hl; c.font = Font(bold=True)
        ws.append([])
        ws.append(["Исходный участок", target])
        ws.append(["Радиус поиска, м", int(RADIUS)])
        ws.append(["Слои", "земельные участки (ВРИ) + здания (назначение); жильё = жилой дом/МКД/ИЖС и т.п."])
        ws.append(["Источник", "Официальный НСПД (nspd.gov.ru) через pynspd, тайловый поиск в контуре"])
        ws.append(["Расстояние", "от границы исходного участка до границы объекта (UTM); касание = 0 м"])
        ws.append(["Оговорка", "Открытые данные НСПД. Для юридически значимого — выписка ЕГРН."])
        for col, w in zip("ABCDEFG", [18, 30, 22, 55, 42, 16, 14]):
            ws.column_dimensions[col].width = w
    wb.save(OUTXLSX)
    wb2 = openpyxl.load_workbook(OUTXLSX)
    return [(sn, wb2[sn].max_row) for sn in wb2.sheetnames]


def main():
    t0 = time.time()
    with npl.Nspd(client_dns_resolve=True, client_retry_on_blocked_ip=True) as nspd:
        ok = False; attempt = 0
        while time.time() - t0 < DEADLINE_MIN * 60:
            attempt += 1
            try:
                if nspd.find_landplot(TARGETS[0][0]) is not None:
                    ok = True; log(f"reachable attempt {attempt} ~{int(time.time()-t0)}s"); break
            except Exception as e:
                log(f"attempt {attempt}: {type(e).__name__}")
            time.sleep(5)
        if not ok:
            json.dump({"ok": False, "error": "timeout"}, open(STATUS, "w", encoding="utf-8"))
            log("TIMEOUT"); return
        per_target = []
        for target, title in TARGETS:
            best, overall = find_for_target(nspd, target)
            per_target.append((target, title, best, overall))
            json.dump({"ok": True, "partial": {t: {"best": b, "overall": ov}
                       for t, _, b, ov in per_target}},
                      open(STATUS, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        verify = write_xlsx(per_target)
        log(f"DONE {OUTXLSX} sheets={verify}")
        print(json.dumps({"ok": True, "verify": verify}, ensure_ascii=False))


if __name__ == "__main__":
    main()
