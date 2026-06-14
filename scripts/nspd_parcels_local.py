#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
nspd_parcels_local.py — COMPLETE land-parcel/ОКС extraction around a target parcel,
straight from the official НСПД via `pynspd` spatial "search in contour".

Unlike the capped public mirror (kadastrmapp api.php = max 81 objects/quarter), this
enumerates EVERY object whose geometry intersects the radius buffer — so every garage.

REQUIRES a Russian (residential) egress IP: НСПД blocks datacenter/foreign IPs. Run it
on the owner's PC with the VPN (AmneziaWG) OFF so the real RU home IP is used. Verify
with:  curl -s -o NUL -w "%{http_code}" https://nspd.gov.ru/   → expect 200, not 000.

    python scripts/nspd_parcels_local.py 18:30:000423:1789 --radius 100 --objects all

Output: an .xlsx next to the script's --out (default: ./<tag>_<radius>m_<cad>.xlsx).
"""
import argparse, json, math, os, sys, time

from shapely.geometry import shape, mapping, box as shp_box
from shapely.ops import transform as shp_transform
from pyproj import Transformer
import openpyxl
from openpyxl.styles import Font

from pynspd import Nspd
from pynspd.schemas import Layer36048Feature, Layer36049Feature  # land parcels / buildings
from pynspd.errors import TooBigContour

SECTORS = ["С", "СВ", "В", "ЮВ", "Ю", "ЮЗ", "З", "СЗ"]


def log(*a):
    print(*a, file=sys.stderr, flush=True)


def geom_of(feat):
    """Return a shapely geometry for a pynspd feature, regardless of accessor."""
    g = getattr(feat, "geometry", None)
    if g is None:
        return None
    # pynspd exposes shapely via .geometry.shape or the geometry is already shapely
    for attr in ("shape", "geom"):
        if hasattr(g, attr):
            try:
                return getattr(g, attr)
            except Exception:
                pass
    try:
        return shape(g if isinstance(g, dict) else g.model_dump())
    except Exception:
        return g


def detect_to_utm(geom):
    """Pick UTM transformer from a geometry that may be in EPSG:3857 or 4326."""
    cx, cy = geom.centroid.x, geom.centroid.y
    if abs(cx) <= 180 and abs(cy) <= 90:          # already lon/lat
        lon, lat, src = cx, cy, 4326
    else:                                          # web mercator
        lon, lat = Transformer.from_crs(3857, 4326, always_xy=True).transform(cx, cy)
        src = 3857
    zone = int((lon + 180) // 6) + 1
    utm = (32600 + zone) if lat >= 0 else (32700 + zone)
    return (Transformer.from_crs(src, utm, always_xy=True).transform,
            Transformer.from_crs(utm, src, always_xy=True).transform, utm, src)


def collect_complete(nspd, layer_def, geom, cad_key, cap=250, depth=0, maxdepth=8, _acc=None):
    """Recursively tile the bbox and gather EVERY feature of a layer.

    pynspd's single intersects call silently caps (~300). We force-split any tile that
    errors with TooBigContour OR returns >= cap features, so nothing is truncated.
    Dedup by cadastral number across overlapping tiles.
    """
    top = _acc is None
    if top:
        _acc = {}
    xmin, ymin, xmax, ymax = geom.bounds
    b = shp_box(xmin, ymin, xmax, ymax)
    too_big = False
    feats = None
    try:
        feats = nspd.search_in_contour(b, layer_def)
    except TooBigContour:
        too_big = True
    if too_big or (feats and len(feats) >= cap and depth < maxdepth):
        mx, my = (xmin + xmax) / 2, (ymin + ymax) / 2
        for q in [(xmin, ymin, mx, my), (mx, my, xmax, ymax),
                  (mx, ymin, xmax, my), (xmin, my, mx, ymax)]:
            collect_complete(nspd, layer_def, shp_box(*q), cad_key, cap, depth + 1, maxdepth, _acc)
    elif feats:
        for f in feats:
            k = cad_key(f)
            if k and k not in _acc:
                _acc[k] = f
    return list(_acc.values()) if top else None


def opt(feat):
    p = getattr(feat, "properties", None)
    o = getattr(p, "options", None) if p else None
    if o is None:
        return {}
    return o if isinstance(o, dict) else (o.model_dump() if hasattr(o, "model_dump") else dict(o))


def bearing_sector(c0, c1):
    dx, dy = c1.x - c0.x, c1.y - c0.y
    ang = (math.degrees(math.atan2(dx, dy)) + 360) % 360
    return SECTORS[int(round(ang / 45.0)) % 8]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("cad_num")
    ap.add_argument("--radius", type=float, default=100.0)
    ap.add_argument("--objects", choices=["land", "all"], default="all",
                    help="land = parcels only; all = + buildings/structures (garages)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--ip-mode", action="store_true",
                    help="hit НСПД by direct IP (no DNS) — use when VPN-off breaks DNS")
    args = ap.parse_args()
    target = args.cad_num.strip()

    nspd_kwargs = {"client_retry_on_blocked_ip": True}
    if args.ip_mode:
        nspd_kwargs["client_dns_resolve"] = True
    with Nspd(**nspd_kwargs) as nspd:
        log(f"== {target} radius {args.radius} m, objects={args.objects} ==")
        central = nspd.find_landplot(target)
        if central is None:
            log("FATAL: target parcel not found (is the RU IP active / VPN off?)"); sys.exit(2)
        cg = geom_of(central)
        to_utm, from_utm, utm, src = detect_to_utm(cg)
        cen_utm = shp_transform(to_utm, cg)
        buf_utm = cen_utm.buffer(args.radius)
        buf_native = shp_transform(from_utm, buf_utm)
        cen_centroid = cen_utm.centroid
        log(f"  central area={opt(central).get('specified_area') or opt(central).get('area')} m^2, "
            f"src=EPSG:{src} UTM=EPSG:{utm}")

        # COMPLETE spatial enumeration inside the buffer (tiled, no silent cap)
        def cad_key(f):
            o = opt(f)
            return o.get("cad_num") or o.get("cad_number") or getattr(f, "cadastral_number", None)

        found = []
        plots = collect_complete(nspd, Layer36048Feature, buf_native, cad_key)
        log(f"  land parcels in contour (complete/tiled): {len(plots)}")
        found += [("Земельный участок", f) for f in plots]
        if args.objects == "all":
            blds = collect_complete(nspd, Layer36049Feature, buf_native, cad_key)
            log(f"  buildings/structures in contour (complete/tiled): {len(blds)}")
            found += [("Здание/сооружение", f) for f in blds]

        rows, seen = [], set()
        for kind, f in found:
            o = opt(f)
            cad = o.get("cad_num") or o.get("cad_number") or getattr(f, "cadastral_number", None)
            if not cad or cad == target or cad in seen:
                continue
            g = geom_of(f)
            if g is None:
                continue
            try:
                g_utm = shp_transform(to_utm, g)
            except Exception:
                continue
            dist = cen_utm.distance(g_utm)
            if dist > args.radius:
                continue
            seen.add(cad)
            rows.append({
                "cad": cad, "kind": kind,
                "addr": o.get("readable_address") or "",
                "vri": (o.get("permitted_use_established_by_document") or o.get("permitted_use")
                        or o.get("land_record_category_type") or o.get("build_record_type_value") or ""),
                "dist": round(dist, 1),
                "side": bearing_sector(cen_centroid, g_utm.centroid),
                "area": o.get("specified_area") or o.get("declared_area") or o.get("area") or "",
                "status": o.get("common_data_status") or o.get("status") or "",
            })
        rows.sort(key=lambda r: r["dist"])
        log(f"  {len(rows)} objects within {args.radius} m of the boundary")

        co = opt(central)
        tag = "vse-obekty" if args.objects == "all" else "uchastki"
        out = args.out or os.path.join(os.getcwd(), f"{tag}_{int(args.radius)}m_{target.replace(':','-')}.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Объекты"
        with_kind = args.objects != "land"
        headers = ["№", "Номер земельного участка", "Адрес земельного участка",
                   "Вид разрешенного использования участка",
                   f"Расстояние от участка {target}, м", "Сторона света",
                   "Площадь, кв. м", "Статус"]
        if with_kind:
            headers.insert(2, "Тип объекта")
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for n, r in enumerate(rows, 1):
            row = [n, r["cad"], r["addr"], r["vri"], r["dist"], r["side"], r["area"], r["status"]]
            if with_kind:
                row.insert(2, r["kind"])
            ws.append(row)
        widths = [5, 22, 60, 40, 28, 14, 14, 16]
        if with_kind:
            widths.insert(2, 20)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        s = wb.create_sheet("Источник")
        for k, v in [
            ("Исходный участок", target),
            ("Адрес", co.get("readable_address", "")),
            ("Площадь, кв. м", co.get("specified_area") or co.get("area") or ""),
            ("Радиус, м", args.radius),
            ("Тип выборки", "все объекты (участки + здания/сооружения)" if with_kind else "земельные участки"),
            ("Найдено объектов", len(rows)),
            ("Источник", "Официальный НСПД (nspd.gov.ru) через pynspd, поиск в контуре"),
            ("Метод", "буфер 100 м от границы исходного участка; расстояние от границы до границы (UTM); касание=0 м"),
            ("Сгенерировано", time.strftime("%Y-%m-%d %H:%M:%S")),
            ("Оговорка", "Открытые данные НСПД. Для юридически значимого использования сверьте с выпиской ЕГРН."),
        ]:
            s.append([k, str(v)])
        s.column_dimensions["A"].width = 22; s.column_dimensions["B"].width = 90
        wb.save(out)

        n = openpyxl.load_workbook(out)["Объекты"].max_row - 1
        log(f"VERIFY: workbook opens, {n} data rows")
        print(json.dumps({"out": out, "objects": len(rows),
                          "central_area": co.get("specified_area") or co.get("area")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
