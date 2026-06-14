#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
nspd_parcels.py — collect ALL Russian cadastral land parcels within a radius of a
target parcel and produce the owner's Excel table (number, address, VRI, distance,
compass side), using the public НСПД/Rosreestr MIRROR (kadastrmapp.online).

WHY THIS EXISTS
---------------
nspd.gov.ru / pkk.rosreestr.ru are IP-blocked for our servers (confirmed 2026-06-14:
http=000 even via the Russian eth0 — НСПД blocks datacenter IPs in general, not just
foreign ones). The ONLY working path from the server is the public mirror, which
proxies НСПД with the right Origin/Referer headers. Do NOT waste time probing the
official map/API directly from the box — it will always time out.

RUN IT WITH THE HERMES VENV PYTHON (it has shapely/pyproj/openpyxl; the terminal
tool's /usr/bin/python3 does NOT):

    /usr/local/lib/hermes-agent/venv/bin/python nspd_parcels.py 18:30:000423:1789 --radius 100

Output: an .xlsx under /root/.hermes/outbox/ (delivered to Telegram as MEDIA:).
"""
import argparse, json, math, os, sys, time, urllib.request, urllib.error

from shapely.geometry import shape, Point
from shapely.ops import transform as shp_transform, unary_union
from pyproj import Transformer
import openpyxl
from openpyxl.styles import Font

MIRRORS = [
    "https://kadastrmapp.online/api.php",
    "https://test.fgishub.ru/api.php",   # fallback; may be down
]
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://kadastrnomer.ru/",
    "Origin": "https://kadastrnomer.ru",
    "Accept": "application/json,*/*",
}
# categoryName buckets returned by the mirror
LAND_HINTS = ("земельн",)                          # "Земельные участки ЕГРН"
OKS_HINTS = ("здани", "сооружени", "незавершён", "незавершен")  # buildings/structures/ОНС (garages live here)
# object-set modes:
#   land = land parcels only (matches the "земельный участок" columns)
#   all  = land + buildings + structures + ОНС (every garage/ОКС, excludes interior "Помещения")
OBJECT_MODES = {
    "land": LAND_HINTS,
    "all": LAND_HINTS + OKS_HINTS,
}
SECTORS = ["С", "СВ", "В", "ЮВ", "Ю", "ЮЗ", "З", "СЗ"]


def log(*a):
    print(*a, file=sys.stderr, flush=True)


def _http_json(url, timeout=25):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.load(r)


def fetch_features(query, timeout=25, retries=2):
    """Query the mirror(s); return list of GeoJSON features (NSPD-style)."""
    from urllib.parse import quote
    last = None
    for base in MIRRORS:
        url = f"{base}?query={quote(query)}&thematicSearchId=1&limit=1000"
        for attempt in range(retries):
            try:
                d = _http_json(url, timeout)
                feats = (d.get("data") or {}).get("features") or d.get("features") or []
                return feats
            except urllib.error.HTTPError as e:
                last = f"HTTP {e.code}"
            except Exception as e:
                last = f"{type(e).__name__}: {str(e)[:60]}"
            time.sleep(0.6)
    log(f"  ! query '{query}' failed via all mirrors ({last})")
    return []


def utm_epsg_for(geom_3857):
    """Pick the right UTM zone (north) from the geometry's longitude."""
    to4326 = Transformer.from_crs(3857, 4326, always_xy=True)
    cx, cy = geom_3857.centroid.x, geom_3857.centroid.y
    lon, lat = to4326.transform(cx, cy)
    zone = int((lon + 180) // 6) + 1
    return (32600 + zone) if lat >= 0 else (32700 + zone)


def bearing_sector(cen_utm_centroid, other_centroid):
    dx = other_centroid.x - cen_utm_centroid.x   # east
    dy = other_centroid.y - cen_utm_centroid.y   # north
    ang = (math.degrees(math.atan2(dx, dy)) + 360) % 360   # 0=N, 90=E
    return SECTORS[int(round(ang / 45.0)) % 8]


def opt(feature):
    return (feature.get("properties") or {}).get("options") or {}


def cat_name(feature):
    return (feature.get("properties") or {}).get("categoryName") or ""


def cat_matches(feature, hints):
    return any(h in cat_name(feature).lower() for h in hints)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("cad_num", help="target cadastral number, e.g. 18:30:000423:1789")
    ap.add_argument("--radius", type=float, default=100.0, help="radius in METERS from target boundary")
    ap.add_argument("--objects", choices=list(OBJECT_MODES), default="land",
                    help="land = land parcels only; all = + buildings/structures/garages (ОКС)")
    ap.add_argument("--quarters-window", type=int, default=8,
                    help="how many numeric neighbour quarters to scan on each side")
    ap.add_argument("--extra-quarters", default="", help="comma-separated extra quarters to scan")
    ap.add_argument("--out", default=None)
    ap.add_argument("--max-cards", type=int, default=600, help="safety cap on per-parcel card fetches")
    args = ap.parse_args()

    target = args.cad_num.strip()
    log(f"== target {target} radius {args.radius} m ==")

    # 1) central parcel polygon
    cfeats = fetch_features(target)
    central = next((f for f in cfeats if opt(f).get("cad_num") == target), cfeats[0] if cfeats else None)
    if not central or not central.get("geometry"):
        log("FATAL: could not fetch target parcel geometry from mirror"); sys.exit(2)
    cen_3857 = shape(central["geometry"])
    if cen_3857.geom_type == "Point":
        log("FATAL: target returned only a Point (no polygon); cannot build a boundary buffer"); sys.exit(2)

    utm = utm_epsg_for(cen_3857)
    to_utm = Transformer.from_crs(3857, utm, always_xy=True).transform
    cen_utm = shp_transform(to_utm, cen_3857)
    cen_centroid = cen_utm.centroid
    buf = cen_utm.buffer(args.radius)
    log(f"  central area={opt(central).get('specified_area') or opt(central).get('area')} m^2, "
        f"UTM=EPSG:{utm}, buffer bbox(m)={[round(v) for v in buf.bounds]}")

    # 2) quarters to scan: home quarter +/- window, plus extras
    parts = target.split(":")
    prefix, qnum = ":".join(parts[:2]), parts[2]
    home_q = f"{prefix}:{qnum}"
    qi = int(qnum)
    width = len(qnum)
    quarters = [f"{prefix}:{str(qi+d).zfill(width)}" for d in range(-args.quarters_window, args.quarters_window + 1)]
    quarters += [q.strip() for q in args.extra_quarters.split(",") if q.strip()]
    seen_q = set(); quarters = [q for q in quarters if not (q in seen_q or seen_q.add(q))]
    log(f"  scanning {len(quarters)} quarters around {home_q}")

    # 3) gather candidate land parcels, pre-filter by representative geometry distance
    candidates = {}   # cad -> feature (from quarter search; may be Point)
    for q in quarters:
        feats = fetch_features(q)
        kept = 0
        hints = OBJECT_MODES[args.objects]
        for f in feats:
            if not cat_matches(f, hints):
                continue
            cad = opt(f).get("cad_num")
            if not cad or cad == target or not f.get("geometry"):
                continue
            try:
                g_utm = shp_transform(to_utm, shape(f["geometry"]))
            except Exception:
                continue
            # pre-filter: point/centroid within radius + 80 m slack of the central buffer
            if cen_utm.distance(g_utm) <= args.radius + 80:
                candidates.setdefault(cad, f); kept += 1
        if feats:
            log(f"    {q}: {len(feats)} objs, {kept} land parcels near buffer")
        time.sleep(0.15)

    log(f"  {len(candidates)} candidate parcels to verify with full polygons")
    if len(candidates) > args.max_cards:
        log(f"  ! capping at {args.max_cards} card fetches");
    # 4) fetch precise polygon per candidate, compute exact distance from boundary
    rows = []
    for i, (cad, qf) in enumerate(sorted(candidates.items())):
        if i >= args.max_cards:
            break
        card = fetch_features(cad)
        cf = next((f for f in card if opt(f).get("cad_num") == cad), card[0] if card else None)
        geom = cf.get("geometry") if cf else qf.get("geometry")
        o = opt(cf) if cf else opt(qf)
        try:
            g_utm = shp_transform(to_utm, shape(geom))
        except Exception:
            continue
        dist = cen_utm.distance(g_utm)
        if dist > args.radius:
            continue
        rows.append({
            "cad": cad,
            "cat": cat_name(cf) if cf else cat_name(qf),
            "addr": o.get("readable_address") or "",
            "vri": (o.get("permitted_use_established_by_document") or o.get("permitted_use")
                    or o.get("land_record_category_type") or o.get("build_record_type_value") or ""),
            "dist": round(dist, 1),
            "side": bearing_sector(cen_centroid, g_utm.centroid),
            "area": o.get("specified_area") or o.get("declared_area") or o.get("area") or "",
            "status": o.get("common_data_status") or o.get("status") or "",
            "pt": g_utm.geom_type,
        })
        time.sleep(0.2)

    rows.sort(key=lambda r: r["dist"])
    log(f"  {len(rows)} parcels within {args.radius} m of the boundary")

    # 5) write workbook
    tag = "vse-obekty" if args.objects == "all" else "uchastki"
    out = args.out or f"/root/.hermes/outbox/{tag}_{int(args.radius)}m_{target.replace(':','-')}.xlsx"
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Участки"
    with_cat = args.objects != "land"
    headers = ["№", "Номер земельного участка", "Адрес земельного участка",
               "Вид разрешенного использования участка",
               f"Расстояние от участка {target}, м", "Сторона света",
               "Площадь, кв. м", "Статус"]
    if with_cat:
        headers.insert(2, "Категория объекта")
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for n, r in enumerate(rows, 1):
        row = [n, r["cad"], r["addr"], r["vri"], r["dist"], r["side"], r["area"], r["status"]]
        if with_cat:
            row.insert(2, r["cat"])
        ws.append(row)
    widths = [5, 22, 60, 40, 28, 14, 14, 16]
    if with_cat:
        widths.insert(2, 22)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    src = wb.create_sheet("Источник")
    co = opt(central)
    for k, v in [
        ("Исходный участок", target),
        ("Адрес", co.get("readable_address", "")),
        ("ВРИ", co.get("permitted_use_established_by_document") or co.get("permitted_use") or ""),
        ("Площадь, кв. м", co.get("specified_area") or co.get("area") or ""),
        ("Радиус, м", args.radius),
        ("Тип выборки", "земельные участки" if args.objects == "land"
                        else "все объекты (участки + здания/сооружения/гаражи-ОКС)"),
        ("Найдено объектов", len(rows)),
        ("Источник данных", "Публичное зеркало НСПД/Росреестр (kadastrmapp.online)"),
        ("Метод расстояния", "от границы исходного участка до границы соседнего (UTM, метры); пересечение/касание = 0 м"),
        ("Сторона света", "по азимуту от центра исходного участка к центру соседнего"),
        ("Сгенерировано", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Оговорка", "Открытые данные публичной кадастровой карты/НСПД. Для юридически значимого использования сверьте с официальной выпиской ЕГРН."),
    ]:
        src.append([k, str(v)])
    src.column_dimensions["A"].width = 24; src.column_dimensions["B"].width = 90
    wb.save(out)

    # 6) verify
    wb2 = openpyxl.load_workbook(out)
    n = wb2["Участки"].max_row - 1
    log(f"VERIFY: workbook opens, {n} data rows")
    print(json.dumps({"out": out, "parcels": len(rows), "candidates": len(candidates),
                      "central_area": co.get("specified_area") or co.get("area")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
